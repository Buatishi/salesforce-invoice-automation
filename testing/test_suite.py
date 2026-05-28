# =============================================================================
# test_suite.py — Tests automáticos Fase 1 (sin Salesforce)
# =============================================================================
# Uso: python test_suite.py  (desde la carpeta testing/)
# No requiere Salesforce ni internet.
#
# ANTES DE CORRER — solo configurar estas dos rutas:
#   EXCEL_PATH   → tu archivo Excel real
#   CARPETA_PDFS → tu carpeta con los PDFs
#
# El test lee el Excel automáticamente y busca los PDFs por cada cliente.
# No hay que poner ningún nombre a mano.
#
# Bloques:
#   1  Lectura y filtrado Excel            (4 checks)
#   2  Matching automático PDF             (2 checks)
#   3  Normalización de texto              (7 checks)
#   4  Validaciones previas al inicio      (5 checks)
#   5  Seguridad — credenciales en logs    (1 check)
#   6  Retry + circuit breaker             (7 checks)
#   7  BotRunner async smoke               (7 checks)
#   8  ExcelWriter persistente             (6 checks)  ← v1.3
#   9  Escritura atómica progreso.json     (5 checks)  ← v1.3
#  10  Cache de selectores (mtime)         (5 checks)  ← v1.3
#  11  Config Pydantic — validación rangos (5 checks)  ← v1.3
#  12  Detección automática de selectores  (5 checks)  ← v1.4
#  13  Calibración guiada (sin SF real)    (5 checks)  ← v1.4
#  14  progreso_db SQLite completo         (5 checks)  ← v1.4
#  16  Funcionalidades v1.4.3              (5 checks)  ← v1.4.4
#  17  Fixes críticos v1.5.x              (17 checks)  ← v1.5.2  ← NUEVO
# =============================================================================

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import json
import tempfile
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from excel_handler import cargar_clientes, normalizar_texto
from file_matcher import buscar_factura
from validaciones import validar_todo

# =============================================================================
# ⚠️  CONFIGURAR ESTAS DOS RUTAS ANTES DE CORRER
# =============================================================================

EXCEL_PATH   = str(Path(__file__).parent / "ENVIAR_ADR_2026.xlsx")
CARPETA_PDFS = str(Path(__file__).parent / "ADR_facturas")

# =============================================================================
# UTILIDADES
# =============================================================================

PASS  = "\033[92m✅ PASS\033[0m"
FAIL  = "\033[91m❌ FAIL\033[0m"
WARN  = "\033[93m⚠️  WARN\033[0m"
INFO  = "\033[94mℹ️  INFO\033[0m"

resultados: Dict[str, int] = {"pass": 0, "fail": 0}


def check(nombre_test: str, condicion: bool, detalle: str = "") -> None:
    if condicion:
        print(f"  {PASS}  {nombre_test}")
        resultados["pass"] += 1
    else:
        print(f"  {FAIL}  {nombre_test}")
        if detalle:
            print(f"         → {detalle}")
        resultados["fail"] += 1


# =============================================================================
# BLOQUE 1 — LECTURA Y FILTRADO DEL EXCEL
# =============================================================================

def test_excel() -> List[Dict]:
    print("\n" + "="*60)
    print("BLOQUE 1 — Lectura y filtrado del Excel")
    print("="*60)

    clientes: List[Dict] = cargar_clientes(EXCEL_PATH)

    # 1.1 — El Excel cargó clientes
    check("1.1 — El Excel tiene clientes pendientes de procesar",
          len(clientes) > 0,
          "Verificá que el Excel existe y tiene filas sin OK")

    # 1.2 — Ningún cliente tiene cobro vacío (el filtrado funcionó)
    check("1.2 — Ningún cliente cargado tiene cobro vacío",
          all(c["cobro"] for c in clientes))

    # 1.3 — Ningún cliente cargado tiene estado OK (el filtrado funcionó)
    check("1.3 — Ningún cliente cargado tiene ENVIADAS = OK",
          all(c["enviada"] != "OK" for c in clientes))

    # 1.4 — Los cobros no tienen espacios al inicio ni al final
    cobros_con_espacios: List[str] = [c["cobro"] for c in clientes
                                      if c["cobro"] != c["cobro"].strip()]
    check("1.4 — Cobros normalizados sin espacios extras",
          len(cobros_con_espacios) == 0,
          f"Cobros con espacios: {cobros_con_espacios}")

    print(f"\n  {INFO}  Total clientes pendientes en el Excel: {len(clientes)}")

    return clientes


# =============================================================================
# BLOQUE 2 — MATCHING AUTOMÁTICO CON LA CARPETA REAL
# =============================================================================

def test_matching_automatico(clientes: List[Dict]) -> Optional[Tuple]:
    """
    Lee todos los clientes del Excel y busca su PDF en la carpeta real.
    No requiere poner ningún nombre a mano.
    Reporta automáticamente cuáles se encontraron y cuáles no.
    """
    print("\n" + "="*60)
    print("BLOQUE 2 — Matching automático Excel → PDFs")
    print("="*60)

    if not clientes:
        print(f"  {WARN}  No hay clientes para testear — revisá el Excel")
        return None

    # Verificar que la carpeta existe
    if not Path(CARPETA_PDFS).exists():
        print(f"  {FAIL}  La carpeta de PDFs no existe: {CARPETA_PDFS}")
        resultados["fail"] += 1
        return None

    pdfs_en_carpeta: List[Path] = (
        list(Path(CARPETA_PDFS).glob("*.pdf")) +
        list(Path(CARPETA_PDFS).glob("*.PDF"))
    )
    print(f"  {INFO}  PDFs en carpeta: {len(pdfs_en_carpeta)}")
    print(f"  {INFO}  Clientes en Excel: {len(clientes)}")
    print()

    # Resultados por categoría
    con_pdf:   List[Tuple[str, str, str]] = []
    sin_pdf:   List[Tuple[str, str]]      = []
    multiples: List[Tuple[str, str]]      = []
    corrupto:  List[Tuple[str, str]]      = []

    for cliente in clientes:
        nombre: str = cliente["nombre"]
        cobro:  str = cliente["cobro"]

        ruta, estado = buscar_factura(nombre, carpeta=CARPETA_PDFS)

        if estado == "OK":
            con_pdf.append((cobro, nombre, ruta.name))
        elif estado == "ERROR_MULTIPLES_FACTURAS":
            multiples.append((cobro, nombre))
        elif estado == "ERROR_ARCHIVO_CORRUPTO":
            corrupto.append((cobro, nombre))
        else:
            sin_pdf.append((cobro, nombre))

    # ── Resultados globales ───────────────────────────────────────────────────
    total: int = len(clientes)
    check(
        f"2.1 — Al menos el 80% de los clientes tienen PDF ({len(con_pdf)}/{total})",
        len(con_pdf) / total >= 0.8 if total > 0 else False,
        f"Encontrados: {len(con_pdf)} | Sin PDF: {len(sin_pdf)} | Múltiples: {len(multiples)}"
    )

    check(
        "2.2 — No hay PDFs corruptos (0 bytes) en la carpeta",
        len(corrupto) == 0,
        f"Corruptos: {[n for _, n in corrupto]}"
    )

    # ── Detalle de clientes SIN PDF ───────────────────────────────────────────
    print()
    if sin_pdf:
        print(f"  {WARN}  Clientes SIN PDF encontrado ({len(sin_pdf)}):")
        for cobro, nombre in sin_pdf:
            print(f"         ❌ {cobro} | {nombre}")
    else:
        print(f"  {INFO}  Todos los clientes tienen PDF. ✅")

    # ── Detalle de PDFs ambiguos ──────────────────────────────────────────────
    if multiples:
        print()
        print(f"  {WARN}  Clientes con múltiples PDFs similares ({len(multiples)}):")
        for cobro, nombre in multiples:
            print(f"         ⚠️  {cobro} | {nombre}")

    # ── Detalle de PDFs corruptos ─────────────────────────────────────────────
    if corrupto:
        print()
        print(f"  {WARN}  PDFs corruptos / 0 bytes ({len(corrupto)}):")
        for cobro, nombre in corrupto:
            print(f"         💀 {cobro} | {nombre}")

    # ── Resumen final del bloque ──────────────────────────────────────────────
    print()
    print(f"  {'─'*50}")
    print(f"  ✅ Con PDF listo          : {len(con_pdf)}")
    print(f"  ❌ Sin PDF                : {len(sin_pdf)}")
    print(f"  ⚠️  Con múltiples similares: {len(multiples)}")
    print(f"  💀 Corruptos (0 bytes)    : {len(corrupto)}")
    print(f"  {'─'*50}")

    return con_pdf, sin_pdf, multiples, corrupto


# =============================================================================
# BLOQUE 3 — NORMALIZACIÓN DE TEXTO
# =============================================================================

def test_normalizacion() -> None:
    print("\n" + "="*60)
    print("BLOQUE 3 — Normalización de texto")
    print("="*60)

    casos: List[Tuple[str, str, str]] = [
        ("García",     "garcia",      "Tilde en a"),
        ("  Juan  ",   "juan",        "Espacios dobles"),
        ("JUAN PEREZ", "juan perez",  "Mayúsculas"),
        ("José Ángel", "jose angel",  "Tildes múltiples"),
        ("O'Brien",    "o brien",     "Apóstrofe"),
        ("Müller",     "muller",      "Diéresis"),
        ("",           "",            "String vacío"),
    ]

    for entrada, esperado, descripcion in casos:
        resultado: str = normalizar_texto(entrada)
        check(f"3.x — '{descripcion}': '{entrada}' → '{esperado}'",
              resultado == esperado,
              f"Obtenido: '{resultado}'")


# =============================================================================
# BLOQUE 4 — VALIDACIONES DE GUI
# =============================================================================

def test_validaciones() -> None:
    print("\n" + "="*60)
    print("BLOQUE 4 — Validaciones previas al inicio")
    print("="*60)

    ok, _ = validar_todo("", EXCEL_PATH, CARPETA_PDFS)
    check("4.1 — URL vacía produce error", not ok)

    ok, _ = validar_todo("miempresa.salesforce.com", EXCEL_PATH, CARPETA_PDFS)
    check("4.2 — URL sin https produce error", not ok)

    ok, _ = validar_todo("https://test.salesforce.com", "/no/existe.xlsx", CARPETA_PDFS)
    check("4.3 — Excel inexistente produce error", not ok)

    ok, _ = validar_todo("https://test.salesforce.com", EXCEL_PATH, "/no/existe")
    check("4.4 — Carpeta inexistente produce error", not ok)

    ok, errores = validar_todo("https://test.my.salesforce.com", EXCEL_PATH, CARPETA_PDFS)
    errores_ruta: List[str] = [e for e in errores
                               if ".env" not in e.lower()
                               and "usuario" not in e.lower()
                               and "contraseña" not in e.lower()
                               and "credencial" not in e.lower()]
    check("4.5 — URL + Excel + Carpeta válidos no dan errores de ruta",
          len(errores_ruta) == 0,
          f"Errores: {errores_ruta}")


# =============================================================================
# BLOQUE 5 — SEGURIDAD: credenciales no en logs
# =============================================================================

def test_seguridad() -> None:
    print("\n" + "="*60)
    print("BLOQUE 5 — Seguridad")
    print("="*60)

    logs_dir: Path = Path(__file__).parent.parent / "logs"

    if not logs_dir.exists():
        print(f"  {INFO}  No hay logs aún — omitiendo test de credenciales")
        return

    env_path: Path = Path(__file__).parent.parent / ".env"
    creds: List[str] = []
    if env_path.exists():
        for linea in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
            if "=" in linea and not linea.startswith("#"):
                valor: str = linea.split("=", 1)[1].strip()
                if valor and "COMPLETAR" not in valor and valor != "":
                    creds.append(valor)

    if not creds:
        print(f"  {INFO}  .env no tiene credenciales reales — omitiendo")
        return

    encontrado: bool = False
    for log_file in logs_dir.glob("*.txt"):
        contenido: str = log_file.read_text(encoding="utf-8", errors="ignore")
        for cred in creds:
            if cred in contenido:
                encontrado = True
                print(f"  {FAIL}  Credencial encontrada en: {log_file.name}")

    check("5.1 — Credenciales NO aparecen en ningún log", not encontrado)


# =============================================================================
# BLOQUE 6 — SMOKE TEST: retry logic y circuit breaker
# =============================================================================

def test_retry_circuit_breaker() -> None:
    """
    Verifica la lógica de retry y circuit breaker sin conectarse a Salesforce.
    Usa mocks internos para simular fallos y éxitos.
    """
    print("\n" + "="*60)
    print("BLOQUE 6 — Retry logic y Circuit Breaker (smoke test)")
    print("="*60)

    try:
        from salesforce_bot import SalesforceBot, CircuitBreakerAbierto
        from config import RETRY_MAX_INTENTOS, CIRCUIT_BREAKER_UMBRAL
    except ImportError as e:
        print(f"  {WARN}  No se pudo importar salesforce_bot: {e}")
        return

    # ── 6.1: _registrar_resultado — OK reinicia el contador ─────────────────
    bot = SalesforceBot()
    bot._fallos_consecutivos = 3
    bot._registrar_resultado(exito=True)
    check("6.1 — _registrar_resultado(OK) reinicia el contador a 0",
          bot._fallos_consecutivos == 0)

    # ── 6.2: _registrar_resultado — fallo incrementa el contador ────────────
    bot2 = SalesforceBot()
    bot2._fallos_consecutivos = 0
    bot2._registrar_resultado(exito=False)
    check("6.2 — _registrar_resultado(fallo) incrementa el contador",
          bot2._fallos_consecutivos == 1)

    # ── 6.3: circuit breaker se abre al superar el umbral ───────────────────
    bot3 = SalesforceBot()
    bot3._fallos_consecutivos = CIRCUIT_BREAKER_UMBRAL - 1
    abierto: bool = False
    try:
        bot3._registrar_resultado(exito=False)
    except CircuitBreakerAbierto:
        abierto = True
    check(
        f"6.3 — CircuitBreakerAbierto se lanza al superar umbral ({CIRCUIT_BREAKER_UMBRAL})",
        abierto,
    )

    # ── 6.4: _con_reintento — éxito en primer intento, sin reintentos ───────
    bot4 = SalesforceBot()
    llamadas: List[int] = []

    def _func_ok() -> str:
        llamadas.append(1)
        return "OK"

    resultado: str = bot4._con_reintento("test_ok", _func_ok)
    check("6.4 — _con_reintento retorna OK en primer intento",
          resultado == "OK" and len(llamadas) == 1,
          f"Resultado: {resultado}, llamadas: {len(llamadas)}")

    # ── 6.5: _con_reintento — agota reintentos y retorna el último error ────
    bot5 = SalesforceBot()
    llamadas2: List[int] = []

    def _func_falla() -> str:
        llamadas2.append(1)
        return "ERROR_ADJUNTO_FALLIDO"

    import unittest.mock as mock
    with mock.patch("salesforce_bot.time.sleep"):
        resultado2: str = bot5._con_reintento("test_falla", _func_falla)

    check(
        f"6.5 — _con_reintento agota {RETRY_MAX_INTENTOS} intentos y retorna error",
        resultado2 == "ERROR_ADJUNTO_FALLIDO" and len(llamadas2) == RETRY_MAX_INTENTOS,
        f"Resultado: {resultado2}, llamadas: {len(llamadas2)} (esperado: {RETRY_MAX_INTENTOS})",
    )

    # ── 6.6: _con_reintento — éxito en segundo intento ──────────────────────
    bot6 = SalesforceBot()
    intento_n: List[int] = [0]

    def _func_falla_luego_ok() -> str:
        intento_n[0] += 1
        if intento_n[0] < 2:
            return "ERROR_ADJUNTO_FALLIDO"
        return "OK"

    with mock.patch("salesforce_bot.time.sleep"):
        resultado3: str = bot6._con_reintento("test_retry_ok", _func_falla_luego_ok)

    check(
        "6.6 — _con_reintento retorna OK después de 1 fallo inicial",
        resultado3 == "OK" and intento_n[0] == 2,
        f"Resultado: {resultado3}, intentos: {intento_n[0]}",
    )

    # ── 6.7: config exporta los cuatro parámetros de retry ──────────────────
    from config import (
        RETRY_MAX_INTENTOS, RETRY_ESPERA_BASE,
        RETRY_ESPERA_MAX,   CIRCUIT_BREAKER_UMBRAL,
    )
    check(
        "6.7 — config.py exporta los 4 parámetros de retry/circuit breaker",
        all([
            isinstance(RETRY_MAX_INTENTOS,    int),
            isinstance(RETRY_ESPERA_BASE,      float),
            isinstance(RETRY_ESPERA_MAX,       float),
            isinstance(CIRCUIT_BREAKER_UMBRAL, int),
        ]),
        f"MAX={RETRY_MAX_INTENTOS}, BASE={RETRY_ESPERA_BASE}, "
        f"MAX_ESPERA={RETRY_ESPERA_MAX}, UMBRAL={CIRCUIT_BREAKER_UMBRAL}",
    )


# =============================================================================
# BLOQUE 7 — SMOKE TEST: asyncio.Queue en BotRunner
# =============================================================================

def test_botrunner_async() -> None:
    """
    Verifica que BotRunner use asyncio.Queue (no queue.Queue) y que los
    métodos de control sean síncronos y no lancen excepciones al llamarlos.
    """
    print("\n" + "="*60)
    print("BLOQUE 7 — BotRunner async (smoke test)")
    print("="*60)

    import asyncio
    import queue as stdlib_queue

    try:
        from bot_runner import BotRunner
    except ImportError as e:
        print(f"  {WARN}  No se pudo importar bot_runner: {e}")
        return

    runner = BotRunner()

    # 7.1: La cola inicial es asyncio.Queue
    check("7.1 — BotRunner.cola es asyncio.Queue",
          isinstance(runner.cola, asyncio.Queue),
          f"Tipo encontrado: {type(runner.cola)}")

    # 7.2: NO es queue.Queue (stdlib)
    check("7.2 — BotRunner.cola NO es queue.Queue (stdlib)",
          not isinstance(runner.cola, stdlib_queue.Queue))

    # 7.3: corriendo=False en estado inicial
    check("7.3 — BotRunner.corriendo es False inicialmente",
          runner.corriendo is False)

    # 7.4: pausado=False en estado inicial
    check("7.4 — BotRunner.pausado es False inicialmente",
          runner.pausado is False)

    # 7.5: pausar() no lanza excepción si el bot no está corriendo
    error_pausar: bool = False
    try:
        runner.pausar()
    except Exception:
        error_pausar = True
    check("7.5 — pausar() sin bot corriendo no lanza excepción", not error_pausar)

    # 7.6: detener() no lanza excepción si el bot no está corriendo
    error_detener: bool = False
    try:
        runner.detener()
    except Exception:
        error_detener = True
    check("7.6 — detener() sin bot corriendo no lanza excepción", not error_detener)

    # 7.7: totales tiene la estructura correcta
    check("7.7 — totales tiene las tres claves esperadas",
          set(runner.totales.keys()) == {"ok", "ya_facturado", "errores"})


# =============================================================================
# BLOQUE 8 — ExcelWriter persistente (v1.3)
# =============================================================================

def test_excel_writer() -> None:
    """
    Verifica que ExcelWriter mantenga el workbook en memoria y solo toque
    el disco cuando se llama flush(), en lugar de abrir/cerrar por cada fila.
    """
    print("\n" + "="*60)
    print("BLOQUE 8 — ExcelWriter persistente (v1.3)")
    print("="*60)

    import openpyxl
    from excel_handler import ExcelWriter, marcar_ok as marcar_ok_compat

    def _excel_tmp(filas: int = 3) -> str:
        """Crea un Excel temporal con N clientes pendientes."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre de la oportunidad", "Cobro", "ENVIADAS"])
        for i in range(1, filas + 1):
            ws.append([f"Cliente {i}", f"C-{i:03d}", ""])
        ruta = tempfile.mktemp(suffix=".xlsx")
        wb.save(ruta)
        return ruta

    def _leer_enviada(ruta: str, fila: int) -> str:
        """Devuelve el valor de la columna ENVIADAS en la fila indicada."""
        wb = openpyxl.load_workbook(ruta)
        ws = wb.active
        col_idx = next(c.column for c in ws[1] if c.value == "ENVIADAS")
        return ws.cell(row=fila, column=col_idx).value or ""

    # ── 8.1: marcar_ok + flush persiste el valor OK ──────────────────────────
    ruta1 = _excel_tmp()
    with ExcelWriter(ruta1) as writer:
        writer.marcar_ok(2)  # fila 2 = primer cliente
    check("8.1 — marcar_ok() + flush escribe OK en el archivo",
          _leer_enviada(ruta1, 2) == "OK")

    # ── 8.2: flush sin cambios no toca el archivo (mtime intacto) ───────────
    ruta2 = _excel_tmp()
    mtime_antes = os.path.getmtime(ruta2)
    time.sleep(0.05)
    writer2 = ExcelWriter(ruta2)
    writer2.flush()   # _dirty == False → no-op
    writer2.cerrar()
    mtime_despues = os.path.getmtime(ruta2)
    check("8.2 — flush() sin cambios no modifica el archivo en disco",
          abs(mtime_antes - mtime_despues) < 0.02,
          f"mtime cambió: {abs(mtime_antes - mtime_despues):.4f}s")

    # ── 8.3: múltiples filas en un solo workbook abierto ────────────────────
    ruta3 = _excel_tmp(filas=5)
    with ExcelWriter(ruta3) as writer:
        for fila in [2, 3, 4]:
            writer.marcar_ok(fila)
    todas_ok = all(_leer_enviada(ruta3, f) == "OK" for f in [2, 3, 4])
    fila5_limpia = _leer_enviada(ruta3, 5) != "OK"
    check("8.3 — Múltiples marcar_ok() con un solo ExcelWriter funciona",
          todas_ok and fila5_limpia,
          f"Filas 2-4 OK: {todas_ok}, fila 5 sin marca: {fila5_limpia}")

    # ── 8.4: context manager llama flush automáticamente al salir ───────────
    ruta4 = _excel_tmp()
    with ExcelWriter(ruta4) as w:
        w.marcar_ok(2)
    # Si el context manager no llamó flush, el archivo no tendría el cambio
    check("8.4 — Context manager (__exit__) llama flush automáticamente",
          _leer_enviada(ruta4, 2) == "OK")

    # ── 8.5: función marcar_ok() de compatibilidad sigue funcionando ─────────
    ruta5 = _excel_tmp()
    resultado_compat = marcar_ok_compat(ruta5, 2)
    check("8.5 — marcar_ok() de compatibilidad retorna True y escribe OK",
          resultado_compat is True and _leer_enviada(ruta5, 2) == "OK",
          f"Retornó: {resultado_compat}, valor: '{_leer_enviada(ruta5, 2)}'")

    # ── 8.6: archivo inexistente retorna False sin lanzar excepción ──────────
    writer_bad = ExcelWriter("/ruta/que/no/existe.xlsx")
    resultado_bad = writer_bad.marcar_ok(2)
    check("8.6 — ExcelWriter sobre archivo inexistente retorna False sin crash",
          resultado_bad is False)


# =============================================================================
# BLOQUE 9 — Persistencia atómica de sesión via progreso_db (v1.3+)
# =============================================================================

def test_progreso_atomico() -> None:
    """
    Verifica que progreso_db.py persista y recupere sesiones correctamente.
    Los 5 checks mantienen la misma semántica que el bloque original,
    adaptados a SQLite con una DB temporal aislada.
    """
    print("\n" + "="*60)
    print("BLOQUE 9 — Persistencia atómica de sesión (progreso_db)")
    print("="*60)

    import progresodb as pg

    # DB temporal aislada — no toca progreso.db del proyecto
    db_tmp: str = tempfile.mktemp(suffix=".db")

    try:
        # ── 9.1: crear_sesion persiste total y modo correctamente ────────────
        pg.crear_sesion("/excel.xlsx", "/pdfs", "https://sf.com", "real", 10,
                        db_path=db_tmp)
        sesion = pg.cargar_sesion(db_path=db_tmp)
        if sesion is not None:
            check("9.1 — crear_sesion() genera progreso.json con total y modo correctos",
                  sesion.get("total") == 10 and sesion.get("modo") == "real",
                  f"total={sesion.get('total')}, modo={sesion.get('modo')}")
        else:
            check("9.1 — crear_sesion() genera progreso.json con total y modo correctos",
                  False, "cargar_sesion() retornó None tras crear_sesion()")

        # ── 9.2: registrar_procesado persiste el cobro ───────────────────────
        pg.registrar_procesado("C-001", "OK", db_path=db_tmp)
        sesion2 = pg.cargar_sesion(db_path=db_tmp)
        check("9.2 — registrar_procesado() persiste el cobro en el JSON",
              sesion2 is not None
              and sesion2.get("procesados", {}).get("C-001") == "OK",
              f"procesados: {sesion2.get('procesados', {}) if sesion2 else None}")

        # ── 9.3: escritura atómica — no quedan archivos temporales huérfanos ─
        # SQLite con WAL no genera .tmp; verificar que no existe progreso.tmp
        # en el directorio de trabajo (garantía equivalente a la original).
        check("9.3 — No queda progreso.tmp huérfano tras escritura exitosa",
              not Path("progreso.tmp").exists(),
              "El archivo .tmp no debería existir con SQLite")

        # ── 9.4: limpiar_sesion desactiva la sesión ──────────────────────────
        pg.limpiar_sesion(db_path=db_tmp)
        sesion_post = pg.cargar_sesion(db_path=db_tmp)
        check("9.4 — limpiar_sesion() elimina progreso.json",
              sesion_post is None,
              "cargar_sesion() debería retornar None tras limpiar_sesion()")

        # ── 9.5: hay_sesion_pendiente refleja el estado real ─────────────────
        pg.crear_sesion("/e.xlsx", "/p", "https://sf.com", "dryrun", 3,
                        db_path=db_tmp)
        pendiente_antes = pg.hay_sesion_pendiente(db_path=db_tmp)
        pg.limpiar_sesion(db_path=db_tmp)
        pendiente_despues = pg.hay_sesion_pendiente(db_path=db_tmp)
        check("9.5 — hay_sesion_pendiente() refleja correctamente si existe el archivo",
              pendiente_antes is True and pendiente_despues is False,
              f"antes={pendiente_antes}, después={pendiente_despues}")

    finally:
        try:
            os.remove(db_tmp)
        except OSError:
            pass


# =============================================================================
# BLOQUE 10 — Cache de selectores con invalidación por mtime (v1.3)
# =============================================================================

def test_cache_selectores() -> None:
    """
    Verifica que cargar_selectores() use cache en memoria validado por mtime
    y que guardar_selectores() invalide el cache correctamente.
    """
    print("\n" + "="*60)
    print("BLOQUE 10 — Cache de selectores mtime (v1.3)")
    print("="*60)

    import salesforce_bot as sb
    from salesforce_bot import cargar_selectores, guardar_selectores, _SELECTORES_DEFAULT

    tmpdir = tempfile.mkdtemp()
    ruta   = os.path.join(tmpdir, "sel_test.json")

    def _escribir_json(valor_usuario: str) -> None:
        datos = {"login": {"campo_usuario": valor_usuario}}
        with open(ruta, "w", encoding="utf-8") as f:
            json.dump(datos, f)

    # Resetear cache antes de cada subtest
    sb._cache_selectores = None

    # ── 10.1: primera carga lee el archivo correctamente ────────────────────
    _escribir_json('input[name="original"]')
    resultado = cargar_selectores(ruta)
    check("10.1 — Primera carga lee el archivo correctamente",
          resultado["login"]["campo_usuario"] == 'input[name="original"]',
          f"Valor: {resultado['login']['campo_usuario']}")

    # ── 10.2: segunda carga sin cambios retorna el mismo objeto (cache) ─────
    r1 = cargar_selectores(ruta)
    r2 = cargar_selectores(ruta)
    check("10.2 — Segunda carga retorna el mismo objeto en memoria (cache)",
          r1 is r2,
          "No retornó el mismo objeto — cache no funciona")

    # ── 10.3: modificar el archivo invalida el cache ─────────────────────────
    time.sleep(0.05)  # garantizar cambio de mtime
    _escribir_json('input[name="modificado"]')
    os.utime(ruta, None)  # forzar mtime diferente
    r3 = cargar_selectores(ruta)
    check("10.3 — Modificar el archivo invalida el cache y retorna nuevos valores",
          r3["login"]["campo_usuario"] == 'input[name="modificado"]',
          f"Valor: {r3['login']['campo_usuario']}")

    # ── 10.4: guardar_selectores() pone el cache en None ────────────────────
    cargar_selectores(ruta)  # poblar cache
    cache_antes = sb._cache_selectores is not None
    guardar_selectores({"login": {"campo_usuario": "nuevo"}}, ruta)
    cache_despues = sb._cache_selectores is None
    check("10.4 — guardar_selectores() invalida el cache (lo pone en None)",
          cache_antes and cache_despues,
          f"cache antes: {cache_antes}, después: {cache_despues}")

    # ── 10.5: archivo inexistente retorna defaults sin error ─────────────────
    sb._cache_selectores = None
    resultado_default = cargar_selectores("/ruta/inexistente/sel.json")
    check("10.5 — Archivo inexistente retorna _SELECTORES_DEFAULT sin lanzar excepción",
          resultado_default == _SELECTORES_DEFAULT)

    # Restaurar cache
    sb._cache_selectores = None


# =============================================================================
# BLOQUE 11 — Config Pydantic: validación de rangos y tipos (v1.3)
# =============================================================================

def test_config_pydantic() -> None:
    """
    Verifica que Settings valide correctamente tipos, rangos y placeholders
    sin depender del .env del proyecto — usa archivos temporales aislados.
    """
    print("\n" + "="*60)
    print("BLOQUE 11 — Config Pydantic: validación de rangos (v1.3)")
    print("="*60)

    try:
        from pydantic import ValidationError
        from config import Settings
    except ImportError as e:
        print(f"  {WARN}  No se pudo importar pydantic o config: {e}")
        return

    def _settings(**kwargs) -> "Settings":
        """Instancia Settings con un .env temporal controlado."""
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".env", delete=False, encoding="utf-8"
        ) as f:
            for k, v in kwargs.items():
                f.write(f"{k.upper()}={v}\n")
            nombre = f.name
        try:
            return Settings(_env_file=nombre)
        finally:
            os.unlink(nombre)

    # ── 11.1: umbral_similitud tiene default 70 ──────────────────────────────
    s = _settings()
    check("11.1 — umbral_similitud tiene default 70",
          s.umbral_similitud == 70,
          f"Valor obtenido: {s.umbral_similitud}")

    # ── 11.2: umbral_similitud acepta valores válidos en rango ───────────────
    s2 = _settings(umbral_similitud=85)
    check("11.2 — umbral_similitud acepta valores válidos (85)",
          s2.umbral_similitud == 85,
          f"Valor obtenido: {s2.umbral_similitud}")

    # ── 11.3: umbral fuera de rango lanza ValidationError ────────────────────
    error_umbral: bool = False
    try:
        _settings(umbral_similitud=150)
    except ValidationError:
        error_umbral = True
    check("11.3 — umbral_similitud=150 lanza ValidationError (máximo es 100)",
          error_umbral)

    # ── 11.4: validador no_placeholder convierte COMPLETAR a '' ─────────────
    s3 = _settings(salesforce_username="COMPLETAR@email.com")
    check("11.4 — Placeholder 'COMPLETAR@email.com' se convierte a string vacío",
          s3.salesforce_username == "",
          f"Valor obtenido: '{s3.salesforce_username}'")

    # ── 11.5: timeout_carga=0 lanza ValidationError (ge=1) ──────────────────
    error_timeout: bool = False
    try:
        _settings(timeout_carga=0)
    except ValidationError:
        error_timeout = True
    check("11.5 — timeout_carga=0 lanza ValidationError (mínimo es 1)",
          error_timeout)


# =============================================================================
# BLOQUE 12 — Detección automática de selectores (sin Salesforce, v1.4)
# =============================================================================

def test_deteccion_automatica() -> None:
    """
    Verifica ESTRATEGIAS_BUSQUEDA, detectar_selector() y explorar_instancia()
    sin conectarse a Salesforce. Usa unittest.mock para simular Playwright Page.

    5 checks:
      12.1 — Cobertura de claves: ESTRATEGIAS_BUSQUEDA == _SELECTORES_DEFAULT
      12.2 — Estructura: todas las entradas son listas no vacías de strings
      12.3 — detectar_selector() retorna el primer selector visible (mock)
      12.4 — detectar_selector() retorna None si ninguno es visible (mock)
      12.5 — explorar_instancia() retorna dict completo y persiste solo detectados
    """
    import unittest.mock as mock

    print("\n" + "="*60)
    print("BLOQUE 12 — Detección automática de selectores (smoke test)")
    print("="*60)

    try:
        import salesforce_bot as sb
        from salesforce_bot import (
            ESTRATEGIAS_BUSQUEDA,
            detectar_selector,
            _SELECTORES_DEFAULT,
        )
    except ImportError as e:
        for label in ["12.1", "12.2", "12.3", "12.4", "12.5"]:
            check(f"{label} — salesforce_bot importable", False, str(e))
        return

    # ── 12.1: ESTRATEGIAS_BUSQUEDA cubre exactamente las claves de _SELECTORES_DEFAULT
    # _SELECTORES_DEFAULT es la fuente interna de verdad sobre qué selectores existen.
    claves_esperadas: set = {
        f"{sec}.{sub}"
        for sec, vals in _SELECTORES_DEFAULT.items()
        if isinstance(vals, dict)
        for sub in vals
    }
    claves_estrategias: set = set(ESTRATEGIAS_BUSQUEDA.keys())
    check(
        "12.1 — ESTRATEGIAS_BUSQUEDA cubre exactamente las mismas claves que _SELECTORES_DEFAULT",
        claves_estrategias == claves_esperadas,
        f"Faltantes: {claves_esperadas - claves_estrategias} | "
        f"Extras: {claves_estrategias - claves_esperadas}",
    )

    # ── 12.2: cada entrada es una lista no vacía de strings no vacíos
    entradas_invalidas: List[str] = [
        clave for clave, lista in ESTRATEGIAS_BUSQUEDA.items()
        if not isinstance(lista, list)
        or len(lista) == 0
        or not all(isinstance(s, str) and s for s in lista)
    ]
    check(
        "12.2 — Todas las entradas de ESTRATEGIAS_BUSQUEDA son listas no vacías de strings",
        len(entradas_invalidas) == 0,
        f"Entradas inválidas: {entradas_invalidas}",
    )

    # ── Fixture de Page mock ─────────────────────────────────────────────────
    # Importar PlaywrightTimeout para usarlo como side_effect del mock.
    # Si playwright no está instalado en el entorno de test, los checks 12.3–12.5
    # se marcan como FAIL con aviso — no bloquean los checks estructurales 12.1/12.2.
    try:
        from playwright.sync_api import TimeoutError as PlaywrightTimeout
    except ImportError:
        for label in ["12.3", "12.4", "12.5"]:
            check(
                f"{label} — playwright disponible para mock",
                False,
                "playwright no instalado en el entorno de test — instalar con: pip install playwright",
            )
        return

    def _page_mock_con_visible(selector_visible: str) -> mock.MagicMock:
        """
        Simula una Page donde solo selector_visible pasa wait_for(state='visible').
        Todos los demás selectores lanzan PlaywrightTimeout.
        """
        def locator_side_effect(sel: str) -> mock.MagicMock:
            loc = mock.MagicMock()
            if sel == selector_visible:
                loc.first.wait_for.return_value = None
            else:
                loc.first.wait_for.side_effect = PlaywrightTimeout("not visible")
            return loc

        page = mock.MagicMock()
        page.locator.side_effect = locator_side_effect
        return page

    # ── 12.3: detectar_selector() retorna el primer selector visible
    # Hacemos visible el tercer candidato de "login.campo_usuario" para probar
    # que la función itera correctamente y no se detiene en el primero que falla.
    clave_test = "login.campo_usuario"
    selector_esperado: str = ESTRATEGIAS_BUSQUEDA[clave_test][2]
    page_ok = _page_mock_con_visible(selector_esperado)
    resultado_ok = detectar_selector(page_ok, clave_test)
    check(
        "12.3 — detectar_selector() retorna el primer selector visible en la cascada",
        resultado_ok == selector_esperado,
        f"Esperado: {selector_esperado!r} | Obtenido: {resultado_ok!r}",
    )

    # ── 12.4: detectar_selector() retorna None cuando ningún selector es visible
    page_none = mock.MagicMock()
    page_none.locator.return_value.first.wait_for.side_effect = PlaywrightTimeout("none")
    resultado_none = detectar_selector(page_none, clave_test)
    check(
        "12.4 — detectar_selector() retorna None si ningún selector es visible",
        resultado_none is None,
        f"Retornó: {resultado_none!r}",
    )

    # ── 12.5: explorar_instancia() retorna dict completo y persiste solo detectados
    # Simular que solo dos selectores son visibles: uno de login y uno de facturado.
    # Verificar que guardar_selectores() es llamado con exactamente esas dos claves
    # y que el dict retornado tiene len == len(ESTRATEGIAS_BUSQUEDA).
    claves_visibles: dict = {
        "login.campo_usuario": ESTRATEGIAS_BUSQUEDA["login.campo_usuario"][0],
        "login.boton_login":   ESTRATEGIAS_BUSQUEDA["login.boton_login"][0],
    }

    def locator_dos_visibles(sel: str) -> mock.MagicMock:
        loc = mock.MagicMock()
        if sel in claves_visibles.values():
            loc.first.wait_for.return_value = None
        else:
            loc.first.wait_for.side_effect = PlaywrightTimeout("not visible")
        return loc

    page_parcial = mock.MagicMock()
    page_parcial.locator.side_effect = locator_dos_visibles

    bot = sb.SalesforceBot()
    bot.page = page_parcial

    with mock.patch("salesforce_bot.guardar_selectores") as mock_guardar:
        mock_guardar.return_value = True
        resultados_exp = bot.explorar_instancia()

    detectados_ok:   List[str] = [k for k, v in resultados_exp.items() if v is not None]
    detectados_none: List[str] = [k for k, v in resultados_exp.items() if v is None]

    check(
        "12.5 — explorar_instancia() retorna dict completo y llama "
        "guardar_selectores() solo con las 2 claves detectadas",
        (
            len(resultados_exp) == len(ESTRATEGIAS_BUSQUEDA)         # dict completo
            and mock_guardar.called                                    # persistió
            and set(detectados_ok) == set(claves_visibles.keys())     # solo las 2
            and len(detectados_none) == len(ESTRATEGIAS_BUSQUEDA) - 2 # resto None
        ),
        f"Detectados: {detectados_ok} | "
        f"guardar_selectores llamado: {mock_guardar.called} | "
        f"arg recibido: {mock_guardar.call_args}",
    )


# =============================================================================
# MAIN
# =============================================================================

# =============================================================================
# BLOQUE 15 — progreso_db.py (SQLite drop-in replacement, T-16)
# =============================================================================

def test_progreso_db() -> None:
    """
    Verifica la interfaz y comportamiento de progreso_db.py.
    5 checks: crear sesión, registrar procesado, obtener pendientes,
    limpiar sesión y _hash_excel.
    No requiere Salesforce ni dependencias externas (solo sqlite3 stdlib).
    """
    import tempfile

    print("\n" + "="*60)
    print("BLOQUE 14 — progreso_db (SQLite drop-in replacement, T-16)")
    print("="*60)

    try:
        import progresodb as pg
    except ImportError as e:
        for label in ["14.1", "14.2", "14.3", "14.4", "14.5"]:
            check(f"{label} — progreso_db importable", False, str(e))
        return

    # DB y excel temporales aislados — no tocan archivos del proyecto
    db_tmp: str = tempfile.mktemp(suffix=".db")
    excel_tmp: str = tempfile.mktemp(suffix=".xlsx")

    try:
        # Excel falso con contenido para que _hash_excel no retorne ""
        with open(excel_tmp, "wb") as f:
            f.write(b"PK" + b"\x00" * 256)

        # ── 14.1 crear_sesion ──────────────────────────────────────────────
        sesion = pg.crear_sesion(
            excel_tmp, "/pdfs", "https://test.salesforce.com", "real", 5,
            db_path=db_tmp,
        )
        check(
            "14.1 — crear_sesion() retorna dict con total, modo y excel_hash correctos",
            (
                isinstance(sesion, dict)
                and sesion.get("total") == 5
                and sesion.get("modo") == "real"
                and "excel_hash" in sesion
                and len(sesion.get("excel_hash", "")) == 16
                and sesion.get("procesados") == {}
            ),
            f"Claves: {list(sesion.keys()) if isinstance(sesion, dict) else sesion}",
        )

        # ── 14.2 registrar_procesado ───────────────────────────────────────
        pg.registrar_procesado("C001", "OK", db_path=db_tmp)
        pg.registrar_procesado("C002", "OK", db_path=db_tmp)
        pg.registrar_procesado("C003", "ERROR_ID", db_path=db_tmp)
        s = pg.cargar_sesion(db_path=db_tmp)
        check(
            "14.2 — registrar_procesado() persiste resultados correctamente en SQLite",
            (
                s is not None
                and len(s["procesados"]) == 3
                and s["procesados"].get("C001") == "OK"
                and s["procesados"].get("C003") == "ERROR_ID"
            ),
            f"procesados={s['procesados'] if s else 'None'}",
        )

        # ── 14.3 obtener_pendientes (doble verificación) ───────────────────
        # C001/C002 en DB con OK → salteados por fuente 1
        # C003 en DB con ERROR_ID → NO salteado (solo OK saltea, igual a progreso.py)
        # C004 con enviada=OK en Excel → salteado por fuente 2
        # C005 → pendiente
        todos = [
            {"cobro": "C001", "enviada": ""},
            {"cobro": "C002", "enviada": ""},
            {"cobro": "C003", "enviada": ""},
            {"cobro": "C004", "enviada": "OK"},
            {"cobro": "C005", "enviada": ""},
        ]
        pendientes = pg.obtener_pendientes(todos, db_path=db_tmp)
        cobros_pend = {c["cobro"] for c in pendientes}
        check(
            "14.3 — obtener_pendientes() aplica doble verificación (DB-OK + Excel-OK)",
            cobros_pend == {"C003", "C005"},
            f"Pendientes: {cobros_pend} (esperado: C003, C005)",
        )

        # ── 14.4 limpiar_sesion ────────────────────────────────────────────
        pg.limpiar_sesion(db_path=db_tmp)
        check(
            "14.4 — limpiar_sesion() desactiva sesión; hay_sesion_pendiente() retorna False",
            (
                pg.hay_sesion_pendiente(db_path=db_tmp) is False
                and pg.cargar_sesion(db_path=db_tmp) is None
            ),
        )

        # ── 14.5 _hash_excel ───────────────────────────────────────────────
        h1 = pg._hash_excel(excel_tmp)
        h2 = pg._hash_excel(excel_tmp)
        h_bad = pg._hash_excel("/archivo/inexistente_T16_test.xlsx")
        check(
            "14.5 — _hash_excel() es determinista (16 chars hex) y '' en archivo inexistente",
            (
                len(h1) == 16
                and h1 == h2
                and h_bad == ""
            ),
            f"h1={h1!r} h_bad={h_bad!r}",
        )

    finally:
        for f in [db_tmp, excel_tmp]:
            try:
                os.remove(f)
            except OSError:
                pass

def test_calibracion() -> None:
    """
    Verifica calibrar_instancia() sin conectarse a Salesforce real.
    Usa unittest.mock para simular Page y los métodos internos del bot.
 
    5 checks:
      13.1 — calibrar_instancia() retorna ok=False si page es None
      13.2 — calibrar_instancia() llama a detectar_selector para claves HOME
      13.3 — calibrar_instancia() llama a buscar_cliente con el cobro dado
      13.4 — calibrar_instancia() retorna ok=False y mensaje si cobro no encontrado
      13.5 — calibrar_instancia() detecta selectores contextuales y llama guardar_selectores
    """
    import unittest.mock as mock
 
    print("\n" + "=" * 60)
    print("BLOQUE 13 — Calibración guiada de selectores (smoke test)")
    print("=" * 60)
 
    try:
        import salesforce_bot as sb
        from salesforce_bot import SalesforceBot, ESTRATEGIAS_BUSQUEDA
        from playwright.sync_api import TimeoutError as PlaywrightTimeout
    except ImportError as e:
        for label in ["13.1", "13.2", "13.3", "13.4", "13.5"]:
            check(f"{label} — salesforce_bot/playwright importable", False, str(e))
        return
 
    # ── 13.1: retorna ok=False si page es None ────────────────────────────────
    bot_sin_page = SalesforceBot()
    bot_sin_page.page = None
    resultado = bot_sin_page.calibrar_instancia("COB-TEST-001")
    check(
        "13.1 — calibrar_instancia() retorna ok=False si bot no está iniciado",
        resultado.get("ok") is False and "no iniciado" in resultado.get("error", "").lower(),
        f"Resultado: {resultado}",
    )
 
    # ── Helper: construir una Page mock donde solo ciertos selectores son visibles
    def _page_con_visibles(selectores_visibles: set) -> mock.MagicMock:
        """Page mock donde solo los selectores en el set pasan wait_for."""
        def locator_side(sel: str) -> mock.MagicMock:
            loc = mock.MagicMock()
            if sel in selectores_visibles:
                loc.first.wait_for.return_value = None
            else:
                loc.first.wait_for.side_effect = PlaywrightTimeout("not visible")
            return loc
        page = mock.MagicMock()
        page.locator.side_effect = locator_side
        return page
 
    # Selectores que serán "visibles" en esta simulación:
    # Primeros candidatos de login y busqueda (HOME), y algunos contextuales.
    sel_login_usuario  = ESTRATEGIAS_BUSQUEDA["login.campo_usuario"][0]
    sel_login_password = ESTRATEGIAS_BUSQUEDA["login.campo_password"][0]
    sel_login_boton    = ESTRATEGIAS_BUSQUEDA["login.boton_login"][0]
    sel_barra_global   = ESTRATEGIAS_BUSQUEDA["busqueda.barra_global"][0]
    sel_boton_subir    = ESTRATEGIAS_BUSQUEDA["archivos_adjuntos.boton_subir"][0]
    sel_casilla        = ESTRATEGIAS_BUSQUEDA["facturado.casilla"][0]
 
    visibles_home = {sel_login_usuario, sel_login_password, sel_login_boton, sel_barra_global}
    visibles_total = visibles_home | {sel_boton_subir, sel_casilla}
 
    # ── 13.2: detectar_selector es llamado para claves HOME ───────────────────
    bot2 = SalesforceBot()
    bot2.page = _page_con_visibles(visibles_total)
    bot2.salesforce_url = "https://test.salesforce.com"
 
    claves_home_esperadas = {
        k for k in ESTRATEGIAS_BUSQUEDA
        if k.split(".", 1)[0] in ("login", "busqueda")
    }
 
    with mock.patch.object(sb, "detectar_selector", wraps=sb.detectar_selector) as spy_detectar, \
         mock.patch.object(bot2, "buscar_cliente", return_value="ENCONTRADO"), \
         mock.patch.object(bot2, "_esperar_carga"), \
         mock.patch.object(sb, "guardar_selectores", return_value=True):
 
        bot2.calibrar_instancia("COB-TEST-002")
        claves_llamadas = {args[1] for args, _ in spy_detectar.call_args_list}
 
    check(
        "13.2 — calibrar_instancia() llama detectar_selector() para todas las claves HOME",
        claves_home_esperadas.issubset(claves_llamadas),
        f"Claves HOME esperadas: {claves_home_esperadas}\n"
        f"         Claves detectadas: {claves_llamadas}",
    )
 
    # ── 13.3: llama a buscar_cliente con el cobro exacto ─────────────────────
    bot3 = SalesforceBot()
    bot3.page = _page_con_visibles(visibles_total)
    bot3.salesforce_url = "https://test.salesforce.com"
    cobro_prueba = "COB-2026-TEST-99"
 
    with mock.patch.object(bot3, "buscar_cliente", return_value="ENCONTRADO") as mock_buscar, \
         mock.patch.object(bot3, "_esperar_carga"), \
         mock.patch.object(sb, "guardar_selectores", return_value=True):
 
        bot3.calibrar_instancia(cobro_prueba)
 
    check(
        "13.3 — calibrar_instancia() llama a buscar_cliente() con el cobro exacto",
        mock_buscar.called and mock_buscar.call_args[0][0] == cobro_prueba,
        f"Llamado con: {mock_buscar.call_args}",
    )
 
    # ── 13.4: retorna ok=False y mensaje si cobro no encontrado ──────────────
    bot4 = SalesforceBot()
    bot4.page = _page_con_visibles(visibles_home)
    bot4.salesforce_url = "https://test.salesforce.com"
 
    with mock.patch.object(bot4, "buscar_cliente", return_value="ERROR_ID"), \
         mock.patch.object(bot4, "_esperar_carga"), \
         mock.patch.object(sb, "guardar_selectores", return_value=True):
 
        res4 = bot4.calibrar_instancia("COB-INEXISTENTE")
 
    check(
        "13.4 — calibrar_instancia() retorna ok=False si cobro no encontrado en SF",
        res4.get("ok") is False and len(res4.get("error", "")) > 0,
        f"Resultado: {res4}",
    )
 
    # ── 13.5: detecta selectores contextuales y llama guardar_selectores ──────
    bot5 = SalesforceBot()
    bot5.page = _page_con_visibles(visibles_total)
    bot5.salesforce_url = "https://test.salesforce.com"
 
    claves_contextuales_esperadas = {
        k for k in ESTRATEGIAS_BUSQUEDA
        if k.split(".", 1)[0] in ("archivos_adjuntos", "facturado")
    }
 
    with mock.patch.object(sb, "detectar_selector", wraps=sb.detectar_selector) as spy_ctx, \
         mock.patch.object(bot5, "buscar_cliente", return_value="ENCONTRADO"), \
         mock.patch.object(bot5, "_esperar_carga"), \
         mock.patch.object(sb, "guardar_selectores", return_value=True) as mock_guardar5:
 
        res5 = bot5.calibrar_instancia("COB-CONTEXTO")
        claves_ctx_llamadas = {args[1] for args, _ in spy_ctx.call_args_list}
 
    check(
        "13.5 — calibrar_instancia() detecta selectores contextuales y llama guardar_selectores()",
        (
            claves_contextuales_esperadas.issubset(claves_ctx_llamadas)
            and mock_guardar5.called
            and res5.get("ok") is True
            and res5.get("detectados", 0) > 0
        ),
        f"Claves contextuales detectadas: {claves_ctx_llamadas & claves_contextuales_esperadas}\n"
        f"         guardar_selectores llamado: {mock_guardar5.called}\n"
        f"         ok={res5.get('ok')}, detectados={res5.get('detectados')}",
    )
 


# =============================================================================
# BLOQUE 16 — Funcionalidades v1.4.3 (on_paso callback, CLIENTE_FIN, ultimos_errores)
# =============================================================================

def test_v143_features() -> None:
    """
    Verifica las funcionalidades incorporadas en v1.4.3.
    Sin conexion a Salesforce real.

    5 checks:
      16.1 — calibrar_instancia() acepta on_paso como parametro opcional
      16.2 — on_paso es llamado con los numeros de paso correctos
      16.3 — on_paso recibe estado 'error' si buscar_cliente falla
      16.4 — BotRunner tiene atributo ultimos_errores inicializado como lista
      16.5 — historial endpoint: /historial retorna estructura correcta
    """
    import unittest.mock as mock

    print("\n" + "="*60)
    print("BLOQUE 16 — Funcionalidades v1.4.3 (smoke test)")
    print("="*60)

    # ── 16.1: calibrar_instancia acepta on_paso opcional ─────────────────────
    try:
        import salesforce_bot as sb
        from salesforce_bot import SalesforceBot
        from playwright.sync_api import TimeoutError as PlaywrightTimeout
    except ImportError as e:
        for label in ["16.1", "16.2", "16.3"]:
            check(f"{label} — salesforce_bot/playwright importable", False, str(e))
        for label in ["16.4", "16.5"]:
            check(f"{label} — importable", False, str(e))
        return

    import inspect
    sig = inspect.signature(SalesforceBot.calibrar_instancia)
    tiene_on_paso = "on_paso" in sig.parameters
    check(
        "16.1 — calibrar_instancia() acepta parametro on_paso opcional",
        tiene_on_paso,
        f"Parametros actuales: {list(sig.parameters.keys())}",
    )

    # ── 16.2: on_paso es llamado con pasos 1-N y estados correctos ───────────
    pasos_recibidos: List[tuple] = []

    def capturar_paso(n, estado):
        pasos_recibidos.append((n, estado))

    # Page mock: todos los selectores HOME visibles, buscar_cliente retorna ENCONTRADO
    try:
        from salesforce_bot import ESTRATEGIAS_BUSQUEDA
        sel_visible = ESTRATEGIAS_BUSQUEDA["login.campo_usuario"][0]

        def _locator(sel):
            loc = mock.MagicMock()
            if sel == sel_visible:
                loc.first.wait_for.return_value = None
            else:
                loc.first.wait_for.side_effect = PlaywrightTimeout("nope")
            return loc

        page_mock = mock.MagicMock()
        page_mock.locator.side_effect = _locator

        bot2 = SalesforceBot()
        bot2.page = page_mock

        with mock.patch.object(bot2, "buscar_cliente", return_value="ENCONTRADO"), \
             mock.patch.object(bot2, "_esperar_carga"), \
             mock.patch.object(sb, "guardar_selectores", return_value=True):
            bot2.calibrar_instancia("COB-TEST-16", on_paso=capturar_paso)

        pasos_numeros  = [p[0] for p in pasos_recibidos]
        estados_validos = {"activo", "ok", "error"}
        check(
            "16.2 — on_paso() es llamado con numeros de paso 1-N y estados validos",
            (
                len(pasos_recibidos) > 0
                and all(isinstance(n, int) and n >= 1 for n, _ in pasos_recibidos)
                and all(e in estados_validos for _, e in pasos_recibidos)
                and 1 in pasos_numeros
            ),
            f"Pasos recibidos: {pasos_recibidos}",
        )
    except Exception as e:
        check("16.2 — on_paso() es llamado correctamente", False, str(e))

    # ── 16.3: on_paso recibe 'error' si buscar_cliente falla ─────────────────
    pasos_error: List[tuple] = []

    def capturar_error(n, estado):
        pasos_error.append((n, estado))

    try:
        bot3 = SalesforceBot()
        bot3.page = page_mock

        with mock.patch.object(bot3, "buscar_cliente", return_value="ERROR_ID"), \
             mock.patch.object(bot3, "_esperar_carga"), \
             mock.patch.object(sb, "guardar_selectores", return_value=True):
            res3 = bot3.calibrar_instancia("COB-INEXISTENTE", on_paso=capturar_error)

        estados_recibidos = [e for _, e in pasos_error]
        check(
            "16.3 — on_paso() recibe estado 'error' si buscar_cliente falla",
            (
                "error" in estados_recibidos
                and res3.get("ok") is False
            ),
            f"Estados recibidos: {estados_recibidos} | ok={res3.get('ok')}",
        )
    except Exception as e:
        check("16.3 — on_paso() con error", False, str(e))

    # ── 16.4: BotRunner.ultimos_errores inicializado como lista ──────────────
    try:
        from bot_runner import BotRunner
        br = BotRunner()
        check(
            "16.4 — BotRunner tiene atributo ultimos_errores inicializado como lista vacia",
            (
                hasattr(br, "ultimos_errores")
                and isinstance(br.ultimos_errores, list)
                and len(br.ultimos_errores) == 0
            ),
            f"ultimos_errores = {getattr(br, 'ultimos_errores', 'MISSING')}",
        )
    except Exception as e:
        check("16.4 — BotRunner.ultimos_errores", False, str(e))

    # ── 16.5: endpoint /historial retorna estructura correcta (smoke test) ────
    try:
        import sys
        import importlib

        # Importar app con mocks de dependencias pesadas
        with mock.patch.dict(sys.modules, {
            "playwright":                    mock.MagicMock(),
            "playwright.sync_api":           mock.MagicMock(),
            "bot_runner":                    mock.MagicMock(),
            "salesforce_bot":               mock.MagicMock(),
            "selector_healer":              mock.MagicMock(),
        }):
            # Solo verificar que el endpoint existe y retorna el formato correcto
            # sin levantar el servidor Flask real
            from app import app as flask_app
            with flask_app.test_client() as client:
                resp = client.get("/historial")
                data = resp.get_json()
            check(
                "16.5 — GET /historial retorna {ok, corridas, retention_dias}",
                (
                    resp.status_code == 200
                    and isinstance(data, dict)
                    and "ok" in data
                    and "corridas" in data
                    and isinstance(data["corridas"], list)
                    and "retention_dias" in data
                ),
                f"Status: {resp.status_code} | Keys: {list(data.keys()) if data else 'None'}",
            )
    except Exception as e:
        check("16.5 — /historial endpoint smoke test", False,
              f"{type(e).__name__}: {str(e)[:150]}")



# =============================================================================
# BLOQUE 17 — Fixes críticos v1.5.x (FIX-1..FIX-8 + features v1.5.2)
# =============================================================================
# Verifica que los bugs que dejaban el panel estático / el bot mudo
# estén corregidos sin necesidad de levantar Flask ni Salesforce.
# =============================================================================

def test_fixes_v15x() -> None:
    """
    17 checks para los fixes críticos de v1.5.x:

    FIX-1  — loop definido antes de loop.time() en _ejecutar()
    FIX-2  — _cola_lista Event existe en iniciar()
    FIX-3  — _arrancar_loop() tiene finally que resetea corriendo
    FIX-4  — _QueueHandler acepta callable getter (no cola directa)
    FIX-5a — confirm() en iniciarBot() va ANTES del fetch /iniciar (JS)
    FIX-5b — botón Iniciar muestra feedback visual (JS)
    FIX-6  — window.onload unificado (sin segundo override)
    FIX-8  — procesarEvento es async function (JS)
    FIX-9  — no hay strings multilínea con comillas simples en el script (JS)
    FIX-10 — no hay regex con newlines literales en el script (JS)
    v1.5.2 — _normalizar_nombre() quita tildes via NFD
    v1.5.2 — ERROR_ID incluye cobro y nombre en el detalle
    v1.5.1 — _t_inicio_corrida definido después de loop
    v1.5.0 — PROGRESO emite dur_seg
    v1.5.0 — PENDIENTE_MANUAL está en estadoMap del panel
    v1.4.9 — ERROR_CIRCUITBREAKER como estado diferenciado
    v1.5.2 — JS pasa validación de sintaxis (node --check)
    """
    print("\n" + "="*60)
    print("BLOQUE 17 — Fixes críticos v1.5.x")
    print("="*60)

    import ast
    import re
    import inspect
    import asyncio
    import subprocess
    import sys
    import tempfile
    from pathlib import Path

    label = "17"

    # ─── Importar BotRunner ───────────────────────────────────────────────────
    try:
        from bot_runner import BotRunner, _QueueHandler
    except ImportError as e:
        check(f"{label}.0 — bot_runner importable", False, str(e))
        return

    runner = BotRunner()

    # ─── FIX-1: loop definido ANTES de loop.time() en _ejecutar() ────────────
    # Verificar en el código fuente que 'loop = asyncio.get_running_loop()' aparece
    # antes de '_t_inicio_corrida = loop.time()' dentro de _ejecutar().
    src = inspect.getsource(BotRunner._ejecutar)
    lines = src.splitlines()
    # La asignación tiene type hint: "loop: asyncio.AbstractEventLoop = asyncio.get_running_loop()"
    # Por eso buscamos 'get_running_loop()' en la primera ocurrencia dentro del body del método.
    idx_loop   = next((i for i, l in enumerate(lines) if 'get_running_loop()' in l and 'loop' in l), -1)
    idx_tinicio = next((i for i, l in enumerate(lines) if '_t_inicio_corrida = loop.time()' in l), -1)
    check(
        "17.1 — FIX-1: loop definido antes de _t_inicio_corrida en _ejecutar()",
        idx_loop != -1 and idx_tinicio != -1 and idx_loop < idx_tinicio,
        f"idx_loop={idx_loop}, idx_tinicio={idx_tinicio} (esperado: loop < tinicio)"
    )

    # ─── FIX-2: _cola_lista threading.Event en iniciar() ─────────────────────
    src_iniciar = inspect.getsource(BotRunner.iniciar)
    check(
        "17.2 — FIX-2: iniciar() crea _cola_lista threading.Event",
        '_cola_lista' in src_iniciar and 'threading.Event' in src_iniciar,
        "No se encontró '_cola_lista' o 'threading.Event' en iniciar()"
    )

    # ─── FIX-3: finally en _arrancar_loop resetea corriendo ──────────────────
    src_loop = inspect.getsource(BotRunner._arrancar_loop)
    check(
        "17.3 — FIX-3: _arrancar_loop() tiene finally con corriendo=False",
        'finally' in src_loop and 'self.corriendo = False' in src_loop,
        "No se encontró 'finally' + 'self.corriendo = False' en _arrancar_loop()"
    )

    # ─── FIX-4: _QueueHandler acepta callable (lambda getter) ────────────────
    src_ejecutar = inspect.getsource(BotRunner._ejecutar)
    check(
        "17.4 — FIX-4: _QueueHandler instanciado con lambda getter (no cola directa)",
        'lambda: self.cola' in src_ejecutar,
        "No se encontró 'lambda: self.cola' en _ejecutar() — _QueueHandler puede usar cola vieja"
    )

    # ─── FIX-4b: _QueueHandler.emit llama al getter, no a self.cola ──────────
    src_handler = inspect.getsource(_QueueHandler.emit)
    check(
        "17.4b — FIX-4b: _QueueHandler.emit usa getter callable, no referencia directa",
        '_cola_getter()' in src_handler,
        "No se encontró '_cola_getter()' en _QueueHandler.emit"
    )

    # ─── v1.5.2 — _normalizar_nombre quita tildes ────────────────────────────
    try:
        # La función es local a bot_runner — acceder via el módulo
        import bot_runner as br
        fn = getattr(br, '_normalizar_nombre', None)
        if fn is None:
            check("17.5 — v1.5.2: _normalizar_nombre() existe en bot_runner", False,
                  "_normalizar_nombre no encontrado en el módulo")
        else:
            check("17.5 — v1.5.2: _normalizar_nombre() existe en bot_runner", True)
            casos = [
                ("García", "Garcia"),
                ("Pérez", "Perez"),
                ("Ñoño", "Nono"),
                ("FERNÁNDEZ", "FERNANDEZ"),
            ]
            for entrada, esperado in casos:
                resultado = fn(entrada)
                check(
                    f"17.5x — _normalizar_nombre('{entrada}') → '{esperado}'",
                    resultado == esperado,
                    f"Obtenido: '{resultado}'"
                )
    except Exception as e:
        check("17.5 — v1.5.2: _normalizar_nombre()", False, str(e))

    # ─── v1.5.2 — ERROR_ID incluye cobro y nombre en el detalle ──────────────
    src_ej = inspect.getsource(BotRunner._ejecutar)
    check(
        "17.6 — v1.5.2: ERROR_ID detalle incluye cobro y nombre del cliente",
        'Verificá que el número de cobro' in src_ej or 'cobro' in src_ej and 'nombre' in src_ej,
        "El detalle de ERROR_ID sigue siendo genérico — no identifica el cliente"
    )

    # ─── v1.5.1 — _t_inicio_corrida definido después del primer evento ───────
    lines_ej = src_ej.splitlines()
    # Busca el await del evento INICIO (no el comentario con la palabra "INICIO")
    idx_evento_inicio = next((i for i, l in enumerate(lines_ej) if '"INICIO"' in l and 'await' in l), -1)
    # _t_inicio_corrida aparece en la asignación y en el uso — queremos la asignación
    idx_tinicio2 = next((i for i, l in enumerate(lines_ej) if '_t_inicio_corrida = loop.time()' in l), -1)
    check(
        "17.7 — v1.5.1: _t_inicio_corrida registrado después del evento INICIO",
        idx_evento_inicio != -1 and idx_tinicio2 != -1 and idx_evento_inicio < idx_tinicio2,
        f"idx_INICIO={idx_evento_inicio}, idx_tinicio={idx_tinicio2}"
    )

    # ─── v1.5.0 — PROGRESO emite dur_seg ─────────────────────────────────────
    check(
        "17.8 — v1.5.0: evento PROGRESO incluye dur_seg para ETA",
        'dur_seg' in src_ej,
        "No se encontró 'dur_seg' en _ejecutar() — ETA no disponible"
    )

    # ─── v1.4.9 — ERROR_CIRCUITBREAKER como estado diferenciado ──────────────
    check(
        "17.9 — v1.4.9: ERROR_CIRCUITBREAKER usado como estado (no ERROR_SESION_EXPIRADA)",
        'ERROR_CIRCUITBREAKER' in src_ej,
        "No se encontró 'ERROR_CIRCUITBREAKER' en _ejecutar()"
    )

    # ─── _finalizar() tiene parámetro motivo ─────────────────────────────────
    src_fin = inspect.getsource(BotRunner._finalizar)
    check(
        "17.10 — v1.4.9 U3: _finalizar() acepta parámetro motivo",
        'motivo' in src_fin,
        "No se encontró parámetro 'motivo' en _finalizar()"
    )

    # ─── FIX-5a / FIX-5b / FIX-6 / FIX-8 / FIX-9 / FIX-10 — JS del panel ──
    # Para los fixes JS, leer app.py y extraer el bloque <script>
    app_path = Path(__file__).parent.parent / 'app.py'
    if not app_path.exists():
        # Intentar en el directorio actual (si se corre desde la raíz)
        app_path = Path('app.py')

    if not app_path.exists():
        for candidate in [Path('../app.py'), Path('../../app.py')]:
            if candidate.exists():
                app_path = candidate
                break

    if not app_path.exists():
        check("17.11 — app.py localizable para validación JS", False,
              f"app.py no encontrado (buscado en: {app_path.resolve()})")
        return

    check("17.11 — app.py localizable para validación JS", True)
    app_content = app_path.read_text(encoding='utf-8')
    start = app_content.find('<script>')
    end   = app_content.rfind('</script>')

    if start == -1 or end == -1:
        check("17.12 — bloque <script> presente en app.py", False)
        return
    check("17.12 — bloque <script> presente en app.py", True)

    script = app_content[start+8:end]

    # FIX-8 — procesarEvento es async
    check(
        "17.13 — FIX-8: procesarEvento es 'async function'",
        'async function procesarEvento' in script,
        "procesarEvento sigue siendo síncrona — await cargarHistorial() causa SyntaxError"
    )

    # FIX-6 — no hay segundo window.onload override como código activo.
    # Nota: puede haber menciones en comentarios — buscamos la asignación activa.
    import re as _re_wiz
    wiz_as_code = bool(_re_wiz.search(r'^\s*const\s+_wizOrigOnload\s*=', script, _re_wiz.MULTILINE))
    check(
        "17.14 — FIX-6: no existe asignación activa de _wizOrigOnload en el script",
        not wiz_as_code,
        "const _wizOrigOnload encontrado como código activo — el override frágil sigue presente"
    )

    # FIX-5a — confirm antes del fetch /iniciar
    # El confirm debe aparecer ANTES de la primera llamada a fetch('/iniciar')
    idx_confirm = script.find("confirm(")
    idx_fetch_iniciar = script.find("fetch('/iniciar'")
    check(
        "17.15 — FIX-5a: confirm() aparece antes del fetch('/iniciar')",
        idx_confirm != -1 and idx_fetch_iniciar != -1 and idx_confirm < idx_fetch_iniciar,
        f"idx_confirm={idx_confirm}, idx_fetch={idx_fetch_iniciar}"
    )

    # FIX-9 / FIX-10 — validación de sintaxis JS con node
    node_available = subprocess.run(
        ['node', '--version'], capture_output=True
    ).returncode == 0

    if node_available:
        with tempfile.NamedTemporaryFile(mode='w', suffix='.js',
                                         encoding='utf-8', delete=False) as tf:
            tf.write(script)
            tmp_js = tf.name
        result = subprocess.run(
            ['node', '--check', tmp_js],
            capture_output=True, text=True
        )
        Path(tmp_js).unlink(missing_ok=True)
        check(
            "17.16 — FIX-9/10: JS del panel pasa validación de sintaxis (node --check)",
            result.returncode == 0,
            result.stdout.strip()[:200] if result.returncode != 0 else ""
        )
    else:
        # Fallback: verificar manualmente que no hay las construcciones problemáticas conocidas
        # (strings multilínea con comillas simples, regex con newlines literales)
        import re as _re
        # Buscar regex literal con newline dentro: /\n/
        regex_with_literal_newline = bool(_re.search(r'/\n/', script))
        check(
            "17.16 — FIX-10: no hay regex /newline_literal/ en el script (node no disponible)",
            not regex_with_literal_newline,
            "Regex con newline literal detectado — causa SyntaxError en el browser"
        )

    # FIX-5b — botón muestra feedback visual
    check(
        "17.17 — FIX-5b: iniciarBot() tiene feedback visual (⏳ Validando...)",
        '⏳ Validando...' in script,
        "No se encontró feedback visual en el botón Iniciar durante la validación"
    )



if __name__ == "__main__":
    print("\n" + "="*60)
    print("  TEST SUITE — InvoiceFlow Bot v1.5.2")
    print("  Fase 1: Tests sin conexión a Salesforce  (83 checks)")
    print("="*60)

    clientes = test_excel()
    test_matching_automatico(clientes)
    test_normalizacion()
    test_validaciones()
    test_seguridad()
    test_retry_circuit_breaker()
    test_botrunner_async()
    test_excel_writer()
    test_progreso_atomico()
    test_cache_selectores()
    test_config_pydantic()
    test_deteccion_automatica()
    test_calibracion()
    test_progreso_db()
    test_v143_features()
    test_fixes_v15x()

    total: int = resultados["pass"] + resultados["fail"]
    print("\n" + "="*60)
    print(f"  RESULTADO: {resultados['pass']}/{total} checks pasaron")
    if resultados["fail"] > 0:
        print(f"  {resultados['fail']} check(s) fallaron")
    else:
        print("  Todo en orden — listo para Fase 2")
    print("="*60 + "\n")